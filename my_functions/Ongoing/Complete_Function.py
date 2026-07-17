#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 28 08:42:24 2019

@author: petern
"""
import shutil
import pandas as pd
from main_def import MAIN_DIR

import os
import pickle
import bcrypt

def get_config():
    """
        How to enscrypt a strings, just for fun here.
        >>> import bcrypt
        >>> password = b"super secret password"
        >>> # Hash a password for the first time, with a randomly-generated salt
        >>> hashed = bcrypt.hashpw(password, bcrypt.gensalt())
        >>> # Check that an unhashed password matches one that has previously been
        >>> # hashed
        >>> if bcrypt.checkpw(password, hashed):
            ...     print("It Matches!")
        ... else:
        ...     print("It Does not Match :(")

        # abc = get_config()
        # print(abc)
    """

    path = os.path.join(MAIN_DIR, "ggl_api")
    path_save = os.path.join(path, "users_config.json")

    try:
        # see if we have run  this before
        import json
        with open(path_save, 'r') as users:
            users_dict = json.load(users)

    except IOError:
        # if not set to default
        print(IOError)

    username = input("Enter your username: ")
    password = input("Enter your password: ")
    get_what = input("What you wanna get: ")
    password = bytes(password, 'utf-8')
    
    hashed = users_dict.get(username)
    if hashed is not None:
        if isinstance(hashed, str):
            hashed = hashed.encode('utf-8')
        
        if bcrypt.checkpw(password, hashed): # unless a password can be None we can use get
            return users_dict.get(get_what)
        else:
            print("Please let it go!")
    else:
        print("Please let it go!")

class list_edit:
    def merge_to_list(*lists):
        '''
        Merge and sort list
        :param lists: <list1>,<list2>
        :return: a sorted list
        Eg: list1 = ["a","b","c"]
            list2 = ["a", "d", "f"]
            list 3 = merge_to_list(list1,list2)
        '''
        print(lists)
        print(type(list))
        newlist = []
        for i in lists:
            newlist.extend(i)
        merge_list = set(newlist)
        merge_list = list(merge_list)
        merge_list.sort()
        return merge_list

class Move_file:
    """
    Move file with ease
    """

    def Data_move(Find_name , Path , Destination_folder):
        """
        Path = "/Users/petern/"
        Destination_folder = "/Users/petern/Desktop/GonJoy/Right Time/OnData/Tool_chat/"
        Find_name = "mytext"
        Move_file.Data_move(Find_name,Find_name, Path, Destination_folder)
        """
        entries = os.listdir(Path)
        for i in entries:
            if pd.Series(i).str.contains(Find_name).bool():
                path = Path + i
                Destination_folder = Destination_folder + i
                shutil.move(path, Destination_folder)

class Time_converse:
    """
    Process time series
    """
    def Time_diff(Segmentation1, id_time, time_col):
        """

        :param Segmentation1:
        :param id_time:
        :param time_col:
        :return:
        """
        Segmentation1['Timediff'] = float()
        Segmentation1[time_col] = pd.to_datetime(Segmentation1[time_col])        
        for i in range(1,Segmentation1.shape[0]):
            if ~pd.isna(Segmentation1[id_time][i]):
                if (Segmentation1[id_time][i] == Segmentation1[id_time][i-1] 
                and Segmentation1[time_col][i].month == Segmentation1[time_col][i-1].month 
                and Segmentation1[time_col][i].day == Segmentation1[time_col][i-1].day 
                and Segmentation1[time_col][i].hour == Segmentation1[time_col][i-1].hour):
                    Segmentation1["Timediff"][i] = (Segmentation1[time_col][i] - Segmentation1[time_col][i-1]).seconds
                elif (Segmentation1[time_col][i].month == Segmentation1[time_col][i-1].month 
                      and Segmentation1[time_col][i].day == Segmentation1[time_col][i-1].day 
                      and Segmentation1[time_col][i].hour == Segmentation1[time_col][i-1].hour):
                    Segmentation1["Timediff"][i-1] = (Segmentation1[time_col][i] - Segmentation1[time_col][i-1]).seconds
                    Segmentation1["Timediff"][i] = (Segmentation1[time_col][i] - Segmentation1[time_col][i-1] ).seconds
                    if (Segmentation1["Timediff"][i-1] == Segmentation1["Timediff"][i]):
                        Segmentation1["Timediff"][i-1] = 0
                else:
                    Segmentation1["Timediff"][i] = 0
            
        return Segmentation1


class Chat_analyse:
    """
        # Connect_final2 = Chat_analyse.teamcheck(Connect_final1, "from_id", "KM_QR", "KM_Content",
        #                                         "khuyen mai|km|khuyến mãi", "ukm|ikm|dkm|kmxpm", Connect_final1)
      Sequence of chat data and create column
    """
    def teamcheck(self, from_id ,AN1_QR, AN1_Content, ANsay, AN_not, Connect_final):
          col1 = [c for c in Connect_final.columns if pd.Series(c).str.contains(from_id).bool()]          
          Connect_final.rename(columns = {col1[0] : "from_id"}, inplace=True)          
          Connect_final =  Connect_final.reset_index(drop = True)
          Connect_final[AN1_QR] =  0
          Connect_final[AN1_Content] =  "0"
          Col_change1 = [c for c in Connect_final.columns if pd.Series(c).str.contains(AN1_QR).bool()]
          Col_change2 = [c for c in Connect_final.columns if pd.Series(c).str.contains(AN1_Content).bool()]          
          MonMansay = ANsay.lower()
          MonMan_not  = AN_not.lower()   
          Connect_final['message'] = Connect_final['message'].fillna("0")                  
          col_mes = [c for c in Connect_final.columns if pd.Series(c).str.contains('message').bool()]
          for k in range(len(Connect_final['from_id'])):          
              An_num = [c for c in Connect_final.loc[k, col_mes].astype(str) if pd.Series(c).str.lower().str.contains(MonMansay).bool()]
              if not An_num == []:  
                  Connect_final.loc[k, Col_change1] = 1
                  Connect_final.loc[k, Col_change2] = Connect_final.loc[k, col_mes].values
          for k in range(len(Connect_final['from_id'])):
              AN_not_num = [c for c in Connect_final.loc[k, col_mes].astype(str) if pd.Series(c).str.lower().str.contains(MonMan_not).bool()]
              if not AN_not_num == []:  
                  Connect_final.loc[k, Col_change1] = 0
                  Connect_final.loc[k, Col_change2] = 0                 
          col_t = [c for c in Connect_final.columns if pd.Series(c).str.contains('from_id').bool()]      
          Connect_final.rename(columns = {col_t[0] : from_id}, 
                     inplace=True)
          return Connect_final


         
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
        # append_df_to_excel("hello.xlsx", pd.DataFrame(np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]]),
        #                columns=['a', 'b', 'c']))
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    """
    import os
    from openpyxl import load_workbook

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    file_exists = os.path.isfile(filename)

    if not file_exists:
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow if startrow is not None else 0, **to_excel_kwargs)
        writer.close()
    else:
        book = load_workbook(filename)
        if startrow is None and sheet_name in book.sheetnames:
            startrow = book[sheet_name].max_row
        elif startrow is None:
            startrow = 0

        if truncate_sheet:
            startrow = 0

        kwargs = {'engine': 'openpyxl', 'mode': 'a'}
        if truncate_sheet:
            kwargs['if_sheet_exists'] = 'replace'
        else:
            kwargs['if_sheet_exists'] = 'overlay'

        with pd.ExcelWriter(filename, **kwargs) as writer:
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, **to_excel_kwargs)

def Extract_Diff_id(Event_Category, Var1, tota_FB_promotion, Question_user):
    tota_FB_promotion = tota_FB_promotion.copy()
    Question_user = Question_user.copy()
    
    col3 = [c for c in Question_user.columns if Var1 in c]
    Question_user.rename(columns={col: "Event.Category" for col in col3}, inplace=True)
    
    col2 = [c for c in tota_FB_promotion.columns if Event_Category in c]
    tota_FB_promotion.rename(columns={col: "Event.Category" for col in col2}, inplace=True)
    
    filter_vals = Question_user["Event.Category"].astype(str).unique()
    tota_FB_promotion = tota_FB_promotion[~tota_FB_promotion["Event.Category"].astype(str).isin(filter_vals)]
    
    tota_FB_promotion.rename(columns={"Event.Category": Event_Category}, inplace=True)
    return tota_FB_promotion

def Same_extract_id(abc_b, abc_c, gonjoybot_chat, Chat_phone_usertx3):
    gonjoybot_chat = gonjoybot_chat.copy()
    Chat_phone_usertx3 = Chat_phone_usertx3.copy()
    
    col3 = [c for c in Chat_phone_usertx3.columns if abc_c in c]
    Chat_phone_usertx3.rename(columns={col: "Event Category" for col in col3}, inplace=True)
    
    col2 = [c for c in gonjoybot_chat.columns if abc_b in c]
    gonjoybot_chat.rename(columns={col: "Event Category" for col in col2}, inplace=True)
    
    keep_vals = Chat_phone_usertx3["Event Category"].astype(str).unique()
    extracted = gonjoybot_chat[gonjoybot_chat["Event Category"].astype(str).isin(keep_vals)].copy()
    
    extracted.rename(columns={"Event Category": abc_b}, inplace=True)
    return extracted

def Get_id_GA_full(Joy_GA, No, DateUpdate):
    import glob
    
    data_path = os.path.expanduser("~/Desktop/GonJoy/Right Time/OnData/Datarequire/")
    files = glob.glob(os.path.join(data_path, "*.csv"))
    
    if not files:
        print("No CSV files found in Datarequire path.")
        return Joy_GA
        
    li = []
    for f in files:
        try:
            li.append(pd.read_csv(f))
        except Exception as e:
            print(f"Error reading {f}: {e}")
            
    if not li:
        return Joy_GA
        
    DatabaseRequire_AllIn = pd.concat(li, axis=0, ignore_index=True)
    
    if 'Date' in DatabaseRequire_AllIn.columns:
        DatabaseRequire_AllIn['date'] = pd.to_datetime(DatabaseRequire_AllIn['Date'].astype(str).str[:8], format='%Y%m%d', errors='coerce')
    else:
        DatabaseRequire_AllIn['date'] = pd.to_datetime(DatabaseRequire_AllIn.iloc[:, 0].astype(str).str[:8], format='%Y%m%d', errors='coerce')
        
    DateUpdate = pd.to_datetime(DateUpdate)
    DatabaseRequire_AllIn = DatabaseRequire_AllIn[DatabaseRequire_AllIn['date'] >= DateUpdate]
    
    def parse_event_cat(val):
        if pd.isna(val):
            return val
        parts = str(val).split('-')
        if len(parts) >= 3:
            res = '-'.join(parts[2:]).strip()
            sub_parts = res.split('-')
            if len(sub_parts) > 1:
                res = '-'.join(sub_parts[1:]).strip()
            return res
        return val

    col_names = list(DatabaseRequire_AllIn.columns)
    if len(col_names) >= 4:
        col3_name = col_names[2]
        col4_name = col_names[3]
        DatabaseRequire_AllIn[col3_name] = DatabaseRequire_AllIn[col4_name].apply(parse_event_cat)

    keep_cols = [c for c in DatabaseRequire_AllIn.columns if 'Event label' in c or 'Event Category' in c]
    DatabaseRequire_AllIn1 = DatabaseRequire_AllIn[keep_cols].copy()
    
    if len(DatabaseRequire_AllIn1.columns) > 0:
        DatabaseRequire_AllIn1.rename(columns={DatabaseRequire_AllIn1.columns[0]: "Var1"}, inplace=True)
        
    Joy_GA = Joy_GA.copy()
    if No < len(Joy_GA.columns):
        Joy_GA.rename(columns={Joy_GA.columns[No]: "Var1"}, inplace=True)
        
    Joy_GA1 = pd.merge(Joy_GA, DatabaseRequire_AllIn1, on="Var1", how="left")
    return Joy_GA1.drop_duplicates()

def Database_Loyalty(loyaltyCustomer, loyaltyProgram, Customer, Provider):
    Customer = Customer.copy()
    Provider = Provider.copy()
    loyaltyProgram = loyaltyProgram.copy()
    loyaltyCustomer = loyaltyCustomer.copy()

    def find_cols(df, pattern):
        return [c for c in df.columns if pattern in str(c)]

    x_id_cus = find_cols(Customer, "X_id")
    created_at_cus = find_cols(Customer, "createdAt")
    if x_id_cus:
        Customer.rename(columns={x_id_cus[0]: "customer"}, inplace=True)
    if created_at_cus:
        Customer.rename(columns={created_at_cus[0]: "createdAt_Cus"}, inplace=True)

    x_id_pro = find_cols(Provider, "X_id")
    created_at_pro = find_cols(Provider, "createdAt")
    if x_id_pro:
        Provider.rename(columns={x_id_pro[0]: "provider"}, inplace=True)
    if created_at_pro:
        Provider.rename(columns={created_at_pro[0]: "createdAt_Pro"}, inplace=True)

    prov_cols = []
    for pat in ["provider", "name", "synonyms", "gps_coordinates", "place"]:
        found = find_cols(Provider, pat)
        if found:
            prov_cols.append(found[0])
    Provider1 = Provider[prov_cols].copy()

    cust_cols = []
    for pat in ["fbId", "name", "createdAt", "lastActiveTime", "customer"]:
        found = find_cols(Customer, pat)
        if found:
            cust_cols.append(found[0])
    if "customer" in Customer.columns and "customer" not in cust_cols:
        cust_cols.append("customer")
    cust_cols = list(dict.fromkeys(cust_cols))
    Customer1 = Customer[cust_cols].copy()

    Loyalty_customer = pd.merge(loyaltyCustomer, Customer1, on="customer", how="left")
    if "fbId" in Loyalty_customer.columns:
        Loyalty_customer["fbId"] = Loyalty_customer["fbId"].astype(str)

    Loyalty_Provider = pd.merge(Loyalty_customer, Provider1, on="provider", how="left")
    
    if "fbId" in Loyalty_Provider.columns:
        Loyalty_Provider = Loyalty_Provider.drop_duplicates(subset=["fbId"])

    prog_cols = []
    for pat in ["X_id", "name", "provider", "description", "levels", "updatedAt"]:
        found = find_cols(loyaltyProgram, pat)
        if found:
            prog_cols.append(found[0])
    
    loyaltyProgram1 = loyaltyProgram[prog_cols].copy()
    up_at_prog = find_cols(loyaltyProgram1, "updatedAt")
    if up_at_prog:
        loyaltyProgram1.rename(columns={up_at_prog[0]: "updatedAt_program"}, inplace=True)

    Loyalty_Provider_Customer_Program = pd.merge(Loyalty_Provider, loyaltyProgram1, on="provider", how="left")
    return Loyalty_Provider_Customer_Program

if __name__ == "__main__":
    import warnings
    warnings.filterwarnings('ignore')
    pd.set_option('float_format', '{:,.2f}'.format)
    pd.set_option("display.max_rows", None, "display.max_columns", 60, 'display.width', 1000)

    data_trial = os.path.join(MAIN_DIR, "data", "Gonj")
    
    gonjoybot_list = sorted(os.listdir(data_trial), reverse=True)[0:2]
    
    li = []
    for filename in gonjoybot_list:
        file_path = os.path.join(data_trial, filename)
        df = pd.read_csv(file_path, index_col=None, header=0)
        li.append(df)
    
    gonjoybot_chat = pd.concat(li, axis=0, ignore_index=True)
    
    gonjoybot_chat['Date'] = pd.to_datetime(gonjoybot_chat['time']) + pd.Timedelta(hours=7)
    gonjoybot_chat = gonjoybot_chat[gonjoybot_chat['Date'].dt.date == pd.to_datetime("2019-06-25").date()]
    
    col1 = [c for c in gonjoybot_chat.columns if pd.Series(c).str.contains('^from|^time$').bool()]
    
    Segmentation = gonjoybot_chat[col1]
    Segmentation1 = Segmentation[~pd.isna(Segmentation["from_id"])].reset_index(drop=True)
    
    Segmentation = Time_converse.Time_diff(Segmentation, "from_id", "time")
    
    print("Time Difference Segmentation:")
    print(Segmentation)
    
    Segmentation = Segmentation.sort_values(by=['time'])
    print("\nSorted by time:")
    print(Segmentation)